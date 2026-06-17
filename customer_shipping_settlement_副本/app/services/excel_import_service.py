from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from app.utils.phone import clean_phone, phone_last4


def parse_customer_workbook(path: Path) -> list:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["拆分结果_加省份"]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        province, city, name, phone_value = row[:4]
        phone = clean_phone(phone_value)
        if not phone or not name or not province:
            continue

        rows.append(
            {
                "province": str(province).strip(),
                "city": "" if city is None else str(city).strip(),
                "name": str(name).strip(),
                "phone": phone,
                "phone_last4": phone_last4(phone),
                "address": "",
                "remark": "",
            }
        )

    return rows


def _number(value: object) -> Optional[float]:
    if value is None or value == "":
        return None
    return float(value)


def parse_price_workbook(path: Path) -> list:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["价格"]
    rows = []

    for row in ws.iter_rows(min_row=6, values_only=True):
        province = row[1]
        if province in ("首重1KG", "二、备注说明"):
            break
        if province is None:
            continue

        rows.append(
            {
                "province": str(province).strip(),
                "price_1kg": _number(row[2]),
                "price_2kg": _number(row[3]),
                "price_3kg": _number(row[4]),
                "price_4kg": _number(row[5]),
                "price_5kg": _number(row[6]),
                "price_6kg": _number(row[7]),
                "over_base_weight_kg": 1.0,
                "over_base_price": _number(row[8]),
                "over_additional_unit_kg": 1.0,
                "over_additional_price": _number(row[9]),
                "remark": "",
            }
        )

    return rows
