import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.settlement_service import get_daily_settlement, get_monthly_settlement


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)


def _prepare_path(output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)


def _write_table(ws, headers: list, rows: list) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 24)


def _shipment_rows(details: list) -> list:
    return [
        [
            row["id"],
            row["ship_date"],
            row["created_at"],
            row["customer_name"],
            row["customer_phone"],
            row["province"],
            row["city"],
            row["address"],
            row["actual_weight_kg"],
            row["billing_weight_kg"],
            row["shipping_fee"],
            row["remark"],
        ]
        for row in details
    ]


def export_customers(conn: sqlite3.Connection, output_path: Path) -> Path:
    rows = conn.execute(
        """
        SELECT id, name, phone, phone_last4, province, city, address,
               is_active, remark, created_at, updated_at
        FROM customers
        ORDER BY id ASC
        """
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "客户资料"
    _write_table(
        ws,
        [
            "客户ID",
            "姓名",
            "电话",
            "电话后4位",
            "省份",
            "城市",
            "地址",
            "状态",
            "备注",
            "创建时间",
            "更新时间",
        ],
        [
            [
                row["id"],
                row["name"],
                row["phone"],
                row["phone_last4"],
                row["province"],
                row["city"],
                row["address"],
                "启用" if row["is_active"] else "停用",
                row["remark"],
                row["created_at"],
                row["updated_at"],
            ]
            for row in rows
        ],
    )
    _prepare_path(output_path)
    wb.save(output_path)
    return output_path


def export_daily_settlement(
    conn: sqlite3.Connection, ship_date: str, output_path: Path
) -> Path:
    settlement = get_daily_settlement(conn, ship_date)
    wb = Workbook()

    overview = wb.active
    overview.title = "每日总览"
    totals = settlement["totals"]
    _write_table(
        overview,
        ["日期", "总件数", "总实际重量", "总结算重量", "总快递费"],
        [
            [
                ship_date,
                totals["shipment_count"],
                totals["total_actual_weight"],
                totals["total_billing_weight"],
                totals["total_fee"],
            ]
        ],
    )

    details = wb.create_sheet("发货明细")
    _write_table(
        details,
        [
            "发货ID",
            "日期",
            "创建时间",
            "姓名",
            "电话",
            "省份",
            "城市",
            "地址",
            "实际重量",
            "结算重量",
            "快递费",
            "备注",
        ],
        _shipment_rows(settlement["details"]),
    )

    province = wb.create_sheet("省份汇总")
    _write_table(
        province,
        ["省份", "件数", "总实际重量", "总结算重量", "总快递费"],
        [
            [
                row["province"],
                row["shipment_count"],
                row["total_actual_weight"],
                row["total_billing_weight"],
                row["total_fee"],
            ]
            for row in settlement["province_summary"]
        ],
    )

    _prepare_path(output_path)
    wb.save(output_path)
    return output_path


def export_monthly_settlement(
    conn: sqlite3.Connection, month: str, output_path: Path
) -> Path:
    settlement = get_monthly_settlement(conn, month)
    wb = Workbook()

    overview = wb.active
    overview.title = "月度总览"
    totals = settlement["totals"]
    _write_table(
        overview,
        ["月份", "总件数", "总实际重量", "总结算重量", "总快递费"],
        [
            [
                month,
                totals["shipment_count"],
                totals["total_actual_weight"],
                totals["total_billing_weight"],
                totals["total_fee"],
            ]
        ],
    )

    date_summary = wb.create_sheet("按日期汇总")
    _write_table(
        date_summary,
        ["日期", "件数", "总实际重量", "总结算重量", "总快递费"],
        [
            [
                row["ship_date"],
                row["shipment_count"],
                row["total_actual_weight"],
                row["total_billing_weight"],
                row["total_fee"],
            ]
            for row in settlement["date_summary"]
        ],
    )

    province = wb.create_sheet("按省份汇总")
    _write_table(
        province,
        ["省份", "件数", "总实际重量", "总结算重量", "总快递费"],
        [
            [
                row["province"],
                row["shipment_count"],
                row["total_actual_weight"],
                row["total_billing_weight"],
                row["total_fee"],
            ]
            for row in settlement["province_summary"]
        ],
    )

    details = wb.create_sheet("发货明细")
    _write_table(
        details,
        [
            "发货ID",
            "日期",
            "创建时间",
            "姓名",
            "电话",
            "省份",
            "城市",
            "地址",
            "实际重量",
            "结算重量",
            "快递费",
            "备注",
        ],
        _shipment_rows(settlement["details"]),
    )

    _prepare_path(output_path)
    wb.save(output_path)
    return output_path
