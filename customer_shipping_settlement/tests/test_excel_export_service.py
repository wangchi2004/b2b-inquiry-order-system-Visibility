from openpyxl import load_workbook

from app.database.connection import connect, initialize_database
from app.services.customer_service import import_customers
from app.services.excel_export_service import (
    export_customers,
    export_daily_settlement,
    export_monthly_settlement,
)
from app.services.price_service import import_prices
from app.services.shipment_service import create_shipment


def seed_export_data(conn):
    import_customers(
        conn,
        [
            {
                "name": "辽宁客户",
                "phone": "18624090227",
                "phone_last4": "0227",
                "province": "辽宁省",
                "city": "沈阳",
                "address": "辽宁地址",
                "remark": "客户备注",
            },
            {
                "name": "吉林客户",
                "phone": "13404421942",
                "phone_last4": "1942",
                "province": "吉林省",
                "city": "镇赉",
                "address": "吉林地址",
                "remark": "",
            },
        ],
        duplicate_mode="update",
    )
    import_prices(
        conn,
        [
            {
                "province": "辽宁省",
                "price_1kg": 2.1,
                "price_2kg": 2.5,
                "price_3kg": 3.0,
                "price_4kg": 4.0,
                "price_5kg": 5.0,
                "price_6kg": 6.5,
                "over_base_weight_kg": 1.0,
                "over_base_price": 3.0,
                "over_additional_unit_kg": 1.0,
                "over_additional_price": 1.0,
                "remark": "",
            },
            {
                "province": "吉林省",
                "price_1kg": 2.1,
                "price_2kg": 2.5,
                "price_3kg": 3.0,
                "price_4kg": 4.0,
                "price_5kg": 5.0,
                "price_6kg": 6.5,
                "over_base_weight_kg": 1.0,
                "over_base_price": 3.0,
                "over_additional_unit_kg": 1.0,
                "over_additional_price": 1.0,
                "remark": "",
            },
        ],
    )
    customers = {
        row["province"]: row["id"]
        for row in conn.execute("SELECT id, province FROM customers").fetchall()
    }
    create_shipment(conn, customers["辽宁省"], 1.5, "2026-06-12", "辽宁当天")
    create_shipment(conn, customers["吉林省"], 5.5, "2026-06-12", "吉林当天")
    create_shipment(conn, customers["辽宁省"], 6.1, "2026-06-13", "辽宁次日")


def test_export_customers_workbook(tmp_path):
    conn = connect(tmp_path / "app.db")
    initialize_database(conn)
    seed_export_data(conn)
    output_path = tmp_path / "customers.xlsx"

    export_customers(conn, output_path)

    wb = load_workbook(output_path, data_only=True)
    ws = wb["客户资料"]
    assert ws["A1"].value == "客户ID"
    assert ws["B1"].value == "姓名"
    assert ws["C1"].value == "电话"
    assert ws["D1"].value == "电话后4位"
    assert ws["B2"].value == "辽宁客户"
    assert ws["C2"].value == "18624090227"
    assert ws["D2"].value == "0227"


def test_export_daily_settlement_workbook(tmp_path):
    conn = connect(tmp_path / "app.db")
    initialize_database(conn)
    seed_export_data(conn)
    output_path = tmp_path / "daily.xlsx"

    export_daily_settlement(conn, "2026-06-12", output_path)

    wb = load_workbook(output_path, data_only=True)
    assert wb.sheetnames == ["每日总览", "发货明细", "省份汇总"]
    overview = wb["每日总览"]
    details = wb["发货明细"]
    province = wb["省份汇总"]
    assert overview["A1"].value == "日期"
    assert overview["B2"].value == 2
    assert overview["E2"].value == 9
    assert details["A1"].value == "发货ID"
    assert details["D2"].value == "辽宁客户"
    assert details["I2"].value == 1.5
    assert details["K2"].value == 2.5
    assert province["A1"].value == "省份"
    assert province["A2"].value == "吉林省"
    assert province["E2"].value == 6.5


def test_export_monthly_settlement_workbook(tmp_path):
    conn = connect(tmp_path / "app.db")
    initialize_database(conn)
    seed_export_data(conn)
    output_path = tmp_path / "monthly.xlsx"

    export_monthly_settlement(conn, "2026-06", output_path)

    wb = load_workbook(output_path, data_only=True)
    assert wb.sheetnames == ["月度总览", "按日期汇总", "按省份汇总", "发货明细"]
    overview = wb["月度总览"]
    date_summary = wb["按日期汇总"]
    province = wb["按省份汇总"]
    assert overview["A1"].value == "月份"
    assert overview["B2"].value == 3
    assert overview["E2"].value == 18
    assert date_summary["A2"].value == "2026-06-12"
    assert date_summary["B2"].value == 2
    assert province["A2"].value == "辽宁省"
    assert province["E2"].value == 11.5
